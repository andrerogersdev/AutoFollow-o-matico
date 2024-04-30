#Bibliotecas
import openpyxl
from time import sleep
from selenium import webdriver
from PySimpleGUI import PySimpleGUI as sg
from selenium.webdriver.common.keys import Keys
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.edge.service import Service as EdgeService

try:
    livro = openpyxl.load_workbook('C:\\Users\\Allya\\OneDrive\\Documentos\\Seguidores\\Planilha Instagram.xlsx')
    Iniciar = True
except FileNotFoundError:
    ColetarDados = True
else:
    ColetarDados = False

if ColetarDados:
    sg.theme('Reddit')
    layout = [

        [sg.Text('-Email ou Nome: \t'), sg.Input(key='Email',size=(30,1))],
        [sg.Text('-Senha: \t\t'),sg.Input(key='senha',password_char='#',size=(30,1))],
        [sg.Button('Confirmar')]

    ]

    janela = sg.Window('InstaBot',layout)


    #ler eventos
    while True:
        eventos,valores = janela.read()
        if eventos == sg.WINDOW_CLOSED:
            Iniciar = False
            break
        if eventos == 'Confirmar':
            Iniciar = True
            email = valores['Email']
            senha = valores['senha']
            book = openpyxl.Workbook()
            book.create_sheet('Instagram')
            InstaPage = book['Instagram']
            InstaPage.append([email,senha])
            book.save('Planilha Instagram.xlsx')
            janela.close()
            break

sleep(10)

while True:
    if Iniciar:
        #Navegador Edge
        Edge_profile = 'user-data-dir=C:\\Users\\Allya\\AppData\\Local\\MicrosoftEdge\\User\\instabot'
        opt = webdriver.EdgeOptions()
        opt.add_argument(Edge_profile)
        opt.add_argument("--window-size=1366,768")
        opt.add_argument('--mute-audio')
        #opt.add_argument('--headless')
        driver = webdriver.Edge(service=EdgeService(EdgeChromiumDriverManager().install()),options=opt)

        driver.get('https://www.instagram.com/')



        def clicar(tipo,valor,valordovalor,escrever=0):
            while True:
                if escrever != 0:
                    try:
                        a = driver.find_element("xpath",f"//{tipo}[contains(@{valor},'{valordovalor}')]")
                        a.click()
                        dormir(3)
                        a.send_keys(escrever)
                        a.send_keys(Keys.ENTER)
                    except:
                        dormir(10)
                    else:
                        print("Clicado com sucesso!")
                        break
                else:
                    try:
                        a = driver.find_element("xpath",f"//{tipo}[contains(@{valor},'{valordovalor}')]")
                        a.click()
                    except:
                        print("não encontrado, esperando 1 segundo.")
                        sleep(1)
                    else:
                        print("clicado com sucesso!")
                        break


        def dormir(tempo):
            print(f"Esperando {tempo} segundos")
            for c in range(0,tempo):
                print(c)
                sleep(1)
            print('-Espera Finalizada-')


        print('Abrindo site')
        dormir(15)
        livro = openpyxl.load_workbook('C:\\Users\\Allya\\OneDrive\\Documentos\\Seguidores\\Planilha Instagram.xlsx')
        Dados = livro['Instagram']
        for rows in Dados.iter_rows(max_row=1):
            email = rows[0].value
            senha = rows[1].value

        def logar(email,senha):
            try:
                logado = driver.find_element("xpath","//img[contains(@crossorigin,'anonymous')]").click()   
            except:
                print("Ainda não logado")

                clicar("input","name","username",escrever=email)
                # Eu poderia apagar as proximas 5 linhas, pois a mesma se resume a linha acima, mas deixarei por orgulho e para que toda vez que eu ver essas linhas resumidas que eu me lembre da minha genialidade

                #logEmail = driver.find_element('xpath','//input[contains(@name,"username")]')
                #dormir(5)
                #logEmail.click()
                #dormir(5)
                #logEmail.send_keys(email)    
                print("Nome inserido com sucesso!") 
                dormir(5)
                clicar("input","name","password",escrever=senha)
                #logSenha = driver.find_element('xpath','//input[contains(@name,"password")]')
                #dormir(5)
                #logSenha.click()
                #logSenha.send_keys(senha)
                #dormir(5)
                #logSenha.send_keys(Keys.ENTER)
                print("senha inserida com sucesso")
                sleep(1)
                print('logando')
            else:
                print("Já logado")
            dormir(20)


        def Seguir():
            driver.get("https://www.instagram.com/explore/people/")
            dormir(10)
            while True:
                seguir = driver.find_elements("xpath","//*[contains(text(), 'Seguir')]")
                try:
                    b = 0
                    for c in range(0,50):
                        seguir[c].click()
                        b += 1
                        if b == 20:
                            print('+20 pessoas seguidas, dormindo 10 minutos')
                            dormir(600)
                            driver.refresh()
                            dormir(15)
                        elif b == 50:
                            reiniciar == True
                            break
                        print("+1 seguido")
                        dormir(5) 
                except:
                    driver.save_screenshot('erro.png')
                    print('erro, reiniciando a página e dormindo 1 hora')
                    dormir(3600)
                    reiniciar = True
                else:
                    reiniciar = False

                if reiniciar:
                    print('reiniciando página')
                    driver.refresh()
                    dormir(10)

        try:
            logar(email,senha)
        except:
            None
        Seguir()
        driver.quit()
        break
    else:
        break